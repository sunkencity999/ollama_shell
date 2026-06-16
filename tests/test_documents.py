"""Tests for the create_document tool (clean successor to file_creation.py)."""

from __future__ import annotations

import pytest

from oshell.providers.base import ToolCall
from oshell.tools import ToolRegistry
from oshell.tools.documents import CreateDocumentTool


def _reg(tmp_path):
    return ToolRegistry([CreateDocumentTool(tmp_path)])


def test_plaintext_and_markdown(tmp_path):
    reg = _reg(tmp_path)
    reg.dispatch(ToolCall(name="create_document", arguments={"path": "a.txt", "content": "hello"}))
    reg.dispatch(ToolCall(name="create_document", arguments={"path": "b.md", "content": "# hi"}))
    assert (tmp_path / "a.txt").read_text() == "hello"
    assert (tmp_path / "b.md").read_text() == "# hi"


def test_csv(tmp_path):
    reg = _reg(tmp_path)
    out = reg.dispatch(
        ToolCall(name="create_document", arguments={"path": "data.csv", "content": "a,b\n1,2"})
    )
    assert not out.startswith("[error]")
    assert (tmp_path / "data.csv").read_text() == "a,b\n1,2"


def test_requires_extension(tmp_path):
    reg = _reg(tmp_path)
    out = reg.dispatch(
        ToolCall(name="create_document", arguments={"path": "noext", "content": "x"})
    )
    assert out.startswith("[error]") and "extension" in out


def test_unsupported_extension(tmp_path):
    reg = _reg(tmp_path)
    out = reg.dispatch(
        ToolCall(name="create_document", arguments={"path": "x.zip", "content": "x"})
    )
    assert out.startswith("[error]") and "unsupported" in out


def test_sandbox_escape_blocked(tmp_path):
    reg = _reg(tmp_path)
    out = reg.dispatch(
        ToolCall(name="create_document", arguments={"path": "../evil.txt", "content": "x"})
    )
    assert out.startswith("[error]") and "escapes" in out


def test_docx(tmp_path):
    pytest.importorskip("docx")
    reg = _reg(tmp_path)
    reg.dispatch(
        ToolCall(
            name="create_document",
            arguments={"path": "doc.docx", "content": "Para one.\n\nPara two."},
        )
    )
    import docx

    paras = [p.text for p in docx.Document(str(tmp_path / "doc.docx")).paragraphs]
    assert "Para one." in paras and "Para two." in paras


def test_xlsx(tmp_path):
    pytest.importorskip("openpyxl")
    reg = _reg(tmp_path)
    reg.dispatch(
        ToolCall(
            name="create_document",
            arguments={"path": "sheet.xlsx", "content": "name,age\nAda,36\nGrace,40"},
        )
    )
    from openpyxl import load_workbook

    ws = load_workbook(str(tmp_path / "sheet.xlsx")).active
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    assert rows[0] == ["name", "age"]
    assert rows[1] == ["Ada", "36"]


def test_pdf(tmp_path):
    pytest.importorskip("weasyprint")
    reg = _reg(tmp_path)
    out = reg.dispatch(
        ToolCall(name="create_document", arguments={"path": "out.pdf", "content": "PDF body"})
    )
    assert not out.startswith("[error]")
    data = (tmp_path / "out.pdf").read_bytes()
    assert data[:4] == b"%PDF"  # valid PDF magic bytes
